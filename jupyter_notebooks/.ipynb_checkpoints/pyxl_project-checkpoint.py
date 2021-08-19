from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# Create a workbook object
#wb = Workbook()

# Load existing spreadsheet
wb = load_workbook('executions.xlsx')
#Create an active worksheet
ws = wb.active
#Print something from ss
print(ws["b2"].value)

